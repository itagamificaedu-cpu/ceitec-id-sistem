"""
views_desempenho.py — Módulo de Análise de Desempenho do Aluno
ITA Tecnologia Educacional / Corretor de Provas (Django)

Perfis suportados e isolamento:
  admin  (ita_admin) → vê todos os resultados, sem filtro
  school (gestor)    → vê apenas professores do seu tenant
  prof   (professor) → vê apenas seus próprios resultados
"""

import json
from collections import defaultdict
from datetime import date
from io import BytesIO
from urllib.parse import quote, unquote

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import DesempenhoAluno, ObservacaoPedagogica, Resultado


# ─── Configurações de classificação ──────────────────────────────────────────

CLASSIFICACAO_INFO = {
    'excelente':          {'label': 'Excelente',           'cor': 'success', 'emoji': '🏆'},
    'satisfatorio':       {'label': 'Satisfatório',         'cor': 'primary', 'emoji': '✅'},
    'em_desenvolvimento': {'label': 'Em Desenvolvimento',   'cor': 'warning', 'emoji': '📈'},
    'critico':            {'label': 'Crítico',              'cor': 'danger',  'emoji': '⚠️'},
}


# ─── Funções auxiliares ───────────────────────────────────────────────────────

def classificar_aluno(media):
    """
    Classifica o aluno pela média:
      Excelente         → média >= 8.5
      Satisfatório      → média >= 6.5
      Em Desenvolvimento→ média >= 4.5
      Crítico           → média < 4.5
    """
    if media >= 8.5:
        return 'excelente'
    elif media >= 6.5:
        return 'satisfatorio'
    elif media >= 4.5:
        return 'em_desenvolvimento'
    return 'critico'


def _get_escola_id(request):
    """Retorna o tenant_id do usuário (usado como escola_id nos registros)."""
    return getattr(request.user, 'tenant_id', None) or 0


def _filtrar_resultados(request):
    """
    Retorna queryset de Resultado filtrado conforme o perfil do usuário,
    garantindo o isolamento completo de dados entre escolas.
    """
    tipo = getattr(request.user, 'type_user', 'prof')
    qs = (
        Resultado.objects
        .filter(nota__isnull=False)
        .select_related('avaliacao', 'avaliacao__professor')
    )
    if tipo == 'admin':
        return qs                                               # sem filtro
    elif tipo == 'school':
        return qs.filter(avaliacao__professor__tenant=request.user.tenant)
    else:
        return qs.filter(avaliacao__professor=request.user)    # próprios


def _calcular_stats_aluno(resultados_lista):
    """
    Recebe uma lista de objetos Resultado e retorna um dicionário de métricas:
      total_avaliacoes, media, maior_nota, menor_nota, taxa_aprovacao,
      evolucao, classificacao, notas_por_disciplina, historico
    Retorna None se a lista estiver vazia.
    """
    notas = [r.nota for r in resultados_lista if r.nota is not None]
    if not notas:
        return None

    total      = len(notas)
    media      = sum(notas) / total
    maior      = max(notas)
    menor      = min(notas)
    aprovadas  = sum(1 for n in notas if n >= 6.0)
    taxa_aprov = round((aprovadas / total) * 100, 1)

    # Evolução: compara última nota com média das 2 primeiras aplicações
    if len(notas) >= 3:
        media_inicio = sum(notas[:2]) / 2
        evolucao = round(notas[-1] - media_inicio, 2)
    else:
        evolucao = 0.0

    # Média por disciplina
    disc_notas = defaultdict(list)
    for r in resultados_lista:
        disc_notas[r.avaliacao.disciplina].append(r.nota)
    notas_por_disciplina = {
        d: round(sum(ns) / len(ns), 2) for d, ns in disc_notas.items()
    }

    # Histórico cronológico de notas
    historico = sorted(
        [
            {
                'data':       (r.avaliacao.data_aplicacao or r.data_correcao.date()).isoformat(),
                'nota':       r.nota,
                'disciplina': r.avaliacao.disciplina,
                'avaliacao':  r.avaliacao.titulo,
            }
            for r in resultados_lista
        ],
        key=lambda x: x['data'],
    )

    return {
        'total_avaliacoes':    total,
        'media':               round(media, 2),
        'maior_nota':          maior,
        'menor_nota':          menor,
        'taxa_aprovacao':      taxa_aprov,
        'evolucao':            evolucao,
        'classificacao':       classificar_aluno(media),
        'notas_por_disciplina': notas_por_disciplina,
        'historico':           historico,
    }


# ─── Views ────────────────────────────────────────────────────────────────────

@login_required
def dashboard_desempenho(request):
    """
    Painel principal de análise de desempenho.
    Exibe resumo por turma: total de alunos, média da turma e distribuição
    de classificações. Inclui dados para Chart.js (gráfico de barras).
    """
    resultados_qs = _filtrar_resultados(request)

    # Agrupa todos os resultados por turma
    grupos_turma = defaultdict(list)
    for r in resultados_qs.order_by('turma', 'aluno_nome'):
        grupos_turma[r.turma].append(r)

    resumo_turmas      = []
    totais_class       = {k: 0 for k in CLASSIFICACAO_INFO}

    for turma, resultados in sorted(grupos_turma.items()):
        alunos_dict = defaultdict(list)
        for r in resultados:
            alunos_dict[r.aluno_nome].append(r)

        medias      = []
        class_count = {k: 0 for k in CLASSIFICACAO_INFO}

        for nome, res in alunos_dict.items():
            stats = _calcular_stats_aluno(res)
            if stats:
                medias.append(stats['media'])
                c = stats['classificacao']
                class_count[c]  += 1
                totais_class[c] += 1

        media_turma = round(sum(medias) / len(medias), 2) if medias else 0

        resumo_turmas.append({
            'turma':          turma,
            'turma_encoded':  quote(turma, safe=''),
            'total_alunos':   len(alunos_dict),
            'media_turma':    media_turma,
            'classificacoes': class_count,
        })

    total_alunos = sum(t['total_alunos'] for t in resumo_turmas)
    media_geral  = 0
    if total_alunos:
        media_geral = round(
            sum(t['media_turma'] * t['total_alunos'] for t in resumo_turmas) / total_alunos, 2
        )

    # Dados para gráfico de barras (médias por turma)
    chart_labels = json.dumps([t['turma'] for t in resumo_turmas])
    chart_medias = json.dumps([t['media_turma'] for t in resumo_turmas])

    return render(request, 'desempenho/dashboard.html', {
        'resumo_turmas':        resumo_turmas,
        'total_turmas':         len(resumo_turmas),
        'total_alunos':         total_alunos,
        'media_geral':          media_geral,
        'totais_classificacao': totais_class,
        'chart_labels':         chart_labels,
        'chart_medias':         chart_medias,
        'classificacao_info':   CLASSIFICACAO_INFO,
    })


@login_required
def turma_resultados(request, turma):
    """
    Lista todos os alunos de uma turma com suas classificações de desempenho.
    Permite busca por nome e exibe gráfico de barras com as médias.
    """
    turma_dec = unquote(turma)
    qs        = _filtrar_resultados(request).filter(turma=turma_dec)

    q_nome = request.GET.get('q', '').strip()
    if q_nome:
        qs = qs.filter(aluno_nome__icontains=q_nome)

    # Agrupa por aluno
    alunos_dict = defaultdict(list)
    for r in qs.order_by('aluno_nome', 'data_correcao'):
        alunos_dict[r.aluno_nome].append(r)

    alunos = []
    for nome, res in alunos_dict.items():
        stats = _calcular_stats_aluno(res)
        if stats:
            info = CLASSIFICACAO_INFO[stats['classificacao']]
            alunos.append({
                'nome':         nome,
                'nome_encoded': quote(nome, safe=''),
                **stats,
                'cor':   info['cor'],
                'label': info['label'],
                'emoji': info['emoji'],
            })

    # Ordena por média decrescente
    alunos.sort(key=lambda a: a['media'], reverse=True)

    medias       = [a['media'] for a in alunos]
    media_turma  = round(sum(medias) / len(medias), 2) if medias else 0

    # Gráfico de barras com primeiro nome apenas (evitar overflow)
    chart_nomes  = json.dumps([a['nome'].split()[0] for a in alunos])
    chart_medias = json.dumps([a['media'] for a in alunos])

    return render(request, 'desempenho/turma.html', {
        'turma':         turma_dec,
        'turma_encoded': quote(turma_dec, safe=''),
        'alunos':        alunos,
        'total_alunos':  len(alunos),
        'media_turma':   media_turma,
        'q_nome':        q_nome,
        'chart_nomes':   chart_nomes,
        'chart_medias':  chart_medias,
        'classificacao_info': CLASSIFICACAO_INFO,
    })


@login_required
def aluno_perfil(request, turma, aluno_nome):
    """
    Perfil detalhado do aluno: evolução de notas (linha), desempenho por
    disciplina (barras) e lista de observações pedagógicas recentes.
    """
    turma_dec = unquote(turma)
    aluno_dec = unquote(aluno_nome)

    resultados = list(
        _filtrar_resultados(request)
        .filter(turma=turma_dec, aluno_nome=aluno_dec)
        .order_by('data_correcao')
    )

    if not resultados:
        messages.warning(
            request,
            f'Nenhum resultado encontrado para {aluno_dec} na turma {turma_dec}.'
        )
        return redirect('turma_resultados', turma=quote(turma_dec, safe=''))

    stats  = _calcular_stats_aluno(resultados)
    info   = CLASSIFICACAO_INFO[stats['classificacao']]
    escola = _get_escola_id(request)
    tipo   = getattr(request.user, 'type_user', 'prof')

    # Observações pedagógicas deste aluno
    obs_qs = ObservacaoPedagogica.objects.filter(
        aluno_nome=aluno_dec, turma=turma_dec
    )
    if escola:
        obs_qs = obs_qs.filter(escola_id=escola)
    elif tipo != 'admin':
        obs_qs = obs_qs.filter(professor=request.user)
    observacoes = obs_qs.order_by('-criado_em')[:10]

    # Dados para Chart.js — evolução (linha)
    hist = stats['historico']
    chart_evo_labels = json.dumps([h['data'] for h in hist])
    chart_evo_notas  = json.dumps([h['nota'] for h in hist])
    chart_evo_disc   = json.dumps([h['disciplina'] for h in hist])

    # Dados para Chart.js — disciplinas (barras)
    disc_keys  = list(stats['notas_por_disciplina'].keys())
    disc_vals  = list(stats['notas_por_disciplina'].values())
    chart_disc_labels = json.dumps(disc_keys)
    chart_disc_notas  = json.dumps(disc_vals)

    # Tabela cronológica de resultados
    tabela = [
        {
            'data':           (r.avaliacao.data_aplicacao or r.data_correcao.date()).isoformat(),
            'disciplina':     r.avaliacao.disciplina,
            'avaliacao':      r.avaliacao.titulo,
            'nota':           r.nota,
            'acertos':        r.acertos,
            'total_questoes': r.avaliacao.numero_questoes,
        }
        for r in resultados
    ]

    return render(request, 'desempenho/aluno.html', {
        'aluno_nome':       aluno_dec,
        'aluno_encoded':    quote(aluno_dec, safe=''),
        'turma':            turma_dec,
        'turma_encoded':    quote(turma_dec, safe=''),
        'stats':            stats,
        'info':             info,
        'observacoes':      observacoes,
        'resultados':       tabela,
        'chart_evo_labels': chart_evo_labels,
        'chart_evo_notas':  chart_evo_notas,
        'chart_evo_disc':   chart_evo_disc,
        'chart_disc_labels': chart_disc_labels,
        'chart_disc_notas':  chart_disc_notas,
        'classificacao_info': CLASSIFICACAO_INFO,
    })


@login_required
def observacoes_lista(request, turma, aluno_nome):
    """Lista todas as observações pedagógicas de um aluno com filtro por tipo."""
    turma_dec = unquote(turma)
    aluno_dec = unquote(aluno_nome)
    escola    = _get_escola_id(request)
    tipo      = getattr(request.user, 'type_user', 'prof')

    obs_qs = ObservacaoPedagogica.objects.filter(
        aluno_nome=aluno_dec, turma=turma_dec
    )
    if escola:
        obs_qs = obs_qs.filter(escola_id=escola)
    elif tipo != 'admin':
        obs_qs = obs_qs.filter(professor=request.user)

    filtro_tipo = request.GET.get('tipo', '')
    if filtro_tipo:
        obs_qs = obs_qs.filter(tipo=filtro_tipo)

    return render(request, 'desempenho/observacoes.html', {
        'aluno_nome':    aluno_dec,
        'aluno_encoded': quote(aluno_dec, safe=''),
        'turma':         turma_dec,
        'turma_encoded': quote(turma_dec, safe=''),
        'observacoes':   obs_qs.order_by('-criado_em'),
        'filtro_tipo':   filtro_tipo,
        'tipos':         ObservacaoPedagogica.TIPOS,
    })


@login_required
def observacao_form(request, turma=None, aluno_nome=None, pk=None):
    """
    Criar uma nova observação pedagógica ou editar uma existente.
    Professor só pode editar as próprias; admin/school edita qualquer uma.
    """
    obs = None
    if pk:
        obs       = get_object_or_404(ObservacaoPedagogica, pk=pk)
        tipo_user = getattr(request.user, 'type_user', 'prof')
        if tipo_user == 'prof' and obs.professor != request.user:
            messages.error(request, 'Você não tem permissão para editar esta observação.')
            return redirect('dashboard_desempenho')
        turma_dec = obs.turma
        aluno_dec = obs.aluno_nome
    else:
        turma_dec = unquote(turma or '')
        aluno_dec = unquote(aluno_nome or '')

    if request.method == 'POST':
        tipo     = request.POST.get('tipo', 'informativo')
        conteudo = request.POST.get('conteudo', '').strip()

        if not conteudo:
            messages.error(request, 'O conteúdo da observação é obrigatório.')
        else:
            if obs:
                obs.tipo     = tipo
                obs.conteudo = conteudo
                obs.save()
                messages.success(request, 'Observação atualizada com sucesso.')
            else:
                ObservacaoPedagogica.objects.create(
                    aluno_nome=aluno_dec,
                    turma=turma_dec,
                    escola_id=_get_escola_id(request),
                    professor=request.user,
                    tipo=tipo,
                    conteudo=conteudo,
                )
                messages.success(request, 'Observação registrada com sucesso.')

            return redirect(
                'aluno_perfil',
                turma=quote(turma_dec, safe=''),
                aluno_nome=quote(aluno_dec, safe=''),
            )

    return render(request, 'desempenho/observacao_form.html', {
        'obs':           obs,
        'aluno_nome':    aluno_dec,
        'aluno_encoded': quote(aluno_dec, safe=''),
        'turma':         turma_dec,
        'turma_encoded': quote(turma_dec, safe=''),
        'tipos':         ObservacaoPedagogica.TIPOS,
    })


@login_required
def observacao_excluir(request, pk):
    """
    Exclui uma observação pedagógica.
    Professor só exclui a própria; admin/school exclui qualquer uma da escola.
    """
    obs       = get_object_or_404(ObservacaoPedagogica, pk=pk)
    tipo_user = getattr(request.user, 'type_user', 'prof')

    if tipo_user == 'prof' and obs.professor != request.user:
        messages.error(request, 'Você não tem permissão para excluir esta observação.')
        return redirect('dashboard_desempenho')

    turma = obs.turma
    aluno = obs.aluno_nome

    if request.method == 'POST':
        obs.delete()
        messages.success(request, 'Observação excluída com sucesso.')
        return redirect(
            'aluno_perfil',
            turma=quote(turma, safe=''),
            aluno_nome=quote(aluno, safe=''),
        )

    return render(request, 'desempenho/observacao_confirmar_exclusao.html', {'obs': obs})


@login_required
def exportar_xlsx(request):
    """
    Exporta planilha XLSX com 5 abas:
      1. Resumo Geral
      2. Por Turma
      3. Por Aluno (classificação + métricas)
      4. Por Disciplina
      5. Observações Pedagógicas
    Aceita filtros opcionais via GET: ?turma=9A&aluno=João
    """
    qs     = _filtrar_resultados(request)
    escola = _get_escola_id(request)
    tipo   = getattr(request.user, 'type_user', 'prof')

    filtro_turma = request.GET.get('turma', '').strip()
    filtro_aluno = request.GET.get('aluno', '').strip()
    if filtro_turma:
        qs = qs.filter(turma=filtro_turma)
    if filtro_aluno:
        qs = qs.filter(aluno_nome__icontains=filtro_aluno)

    resultados = list(qs.order_by('turma', 'aluno_nome', 'data_correcao'))

    # Observações
    obs_qs = ObservacaoPedagogica.objects.all()
    if escola:
        obs_qs = obs_qs.filter(escola_id=escola)
    elif tipo != 'admin':
        obs_qs = obs_qs.filter(professor=request.user)
    if filtro_turma:
        obs_qs = obs_qs.filter(turma=filtro_turma)
    observacoes = list(obs_qs.order_by('turma', 'aluno_nome', '-criado_em'))

    # ─── Workbook ─────────────────────────────────────────────────
    wb = Workbook()

    # Estilos reutilizáveis
    fill_header = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    fill_alt    = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    f_titulo    = Font(bold=True, color='FFFFFF', size=11)
    f_negrito   = Font(bold=True, size=11)
    f_normal    = Font(size=11)
    al_centro   = Alignment(horizontal='center', vertical='center')

    def cabecalho(ws, colunas):
        for col, titulo in enumerate(colunas, 1):
            c = ws.cell(row=1, column=col, value=titulo)
            c.font      = f_titulo
            c.fill      = fill_header
            c.alignment = al_centro

    def ajustar(ws):
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 42)

    # ── Aba 1: Resumo Geral ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumo Geral'
    total_notas = [r.nota for r in resultados if r.nota is not None]
    cabecalho(ws1, ['Métrica', 'Valor'])
    dados = [
        ('Total de Resultados', len(resultados)),
        ('Alunos Únicos', len({r.aluno_nome for r in resultados})),
        ('Turmas', len({r.turma for r in resultados})),
        ('Média Geral', round(sum(total_notas) / len(total_notas), 2) if total_notas else 0),
        ('Maior Nota', max(total_notas, default=0)),
        ('Menor Nota', min(total_notas, default=0)),
        ('Taxa de Aprovação (%)',
         round(sum(1 for n in total_notas if n >= 6) / len(total_notas) * 100, 1)
         if total_notas else 0),
        ('Data de Exportação', date.today().isoformat()),
    ]
    for i, (k, v) in enumerate(dados, 2):
        ws1.cell(row=i, column=1, value=k).font = f_negrito
        ws1.cell(row=i, column=2, value=v).font = f_normal
        if i % 2 == 0:
            ws1.cell(row=i, column=1).fill = fill_alt
            ws1.cell(row=i, column=2).fill = fill_alt
    ajustar(ws1)

    # ── Aba 2: Por Turma ─────────────────────────────────────────
    ws2 = wb.create_sheet('Por Turma')
    cabecalho(ws2, ['Turma', 'Alunos', 'Média', 'Maior Nota', 'Menor Nota', 'Aprovação (%)'])
    grupos_t = defaultdict(list)
    for r in resultados:
        grupos_t[r.turma].append(r)
    for i, (turma, res) in enumerate(sorted(grupos_t.items()), 2):
        notas_t = [r.nota for r in res if r.nota is not None]
        ws2.cell(row=i, column=1, value=turma)
        ws2.cell(row=i, column=2, value=len({r.aluno_nome for r in res}))
        ws2.cell(row=i, column=3, value=round(sum(notas_t) / len(notas_t), 2) if notas_t else 0)
        ws2.cell(row=i, column=4, value=max(notas_t, default=0))
        ws2.cell(row=i, column=5, value=min(notas_t, default=0))
        ws2.cell(row=i, column=6,
                 value=round(sum(1 for n in notas_t if n >= 6) / len(notas_t) * 100, 1)
                 if notas_t else 0)
        if i % 2 == 0:
            for c in range(1, 7):
                ws2.cell(row=i, column=c).fill = fill_alt
    ajustar(ws2)

    # ── Aba 3: Por Aluno ─────────────────────────────────────────
    ws3 = wb.create_sheet('Por Aluno')
    cabecalho(ws3, ['Turma', 'Aluno', 'Avaliações', 'Média',
                    'Maior', 'Menor', 'Aprovação (%)', 'Classificação'])
    grupos_a = defaultdict(list)
    for r in resultados:
        grupos_a[(r.turma, r.aluno_nome)].append(r)
    linha = 2
    for (turma, nome), res in sorted(grupos_a.items()):
        st = _calcular_stats_aluno(res)
        if not st:
            continue
        ws3.cell(row=linha, column=1, value=turma)
        ws3.cell(row=linha, column=2, value=nome)
        ws3.cell(row=linha, column=3, value=st['total_avaliacoes'])
        ws3.cell(row=linha, column=4, value=st['media'])
        ws3.cell(row=linha, column=5, value=st['maior_nota'])
        ws3.cell(row=linha, column=6, value=st['menor_nota'])
        ws3.cell(row=linha, column=7, value=st['taxa_aprovacao'])
        ws3.cell(row=linha, column=8, value=CLASSIFICACAO_INFO[st['classificacao']]['label'])
        if linha % 2 == 0:
            for c in range(1, 9):
                ws3.cell(row=linha, column=c).fill = fill_alt
        linha += 1
    ajustar(ws3)

    # ── Aba 4: Por Disciplina ────────────────────────────────────
    ws4 = wb.create_sheet('Por Disciplina')
    cabecalho(ws4, ['Disciplina', 'Aplicações', 'Média', 'Maior Nota', 'Menor Nota'])
    grupos_d = defaultdict(list)
    for r in resultados:
        grupos_d[r.avaliacao.disciplina].append(r.nota)
    for i, (disc, ns) in enumerate(sorted(grupos_d.items()), 2):
        ns = [n for n in ns if n is not None]
        ws4.cell(row=i, column=1, value=disc)
        ws4.cell(row=i, column=2, value=len(ns))
        ws4.cell(row=i, column=3, value=round(sum(ns) / len(ns), 2) if ns else 0)
        ws4.cell(row=i, column=4, value=max(ns, default=0))
        ws4.cell(row=i, column=5, value=min(ns, default=0))
        if i % 2 == 0:
            for c in range(1, 6):
                ws4.cell(row=i, column=c).fill = fill_alt
    ajustar(ws4)

    # ── Aba 5: Observações ───────────────────────────────────────
    ws5 = wb.create_sheet('Observações')
    cabecalho(ws5, ['Turma', 'Aluno', 'Tipo', 'Observação', 'Professor', 'Data'])
    for i, obs in enumerate(observacoes, 2):
        ws5.cell(row=i, column=1, value=obs.turma)
        ws5.cell(row=i, column=2, value=obs.aluno_nome)
        ws5.cell(row=i, column=3, value=obs.get_tipo_display())
        ws5.cell(row=i, column=4, value=obs.conteudo)
        ws5.cell(row=i, column=5, value=obs.professor.get_full_name() or obs.professor.username)
        ws5.cell(row=i, column=6, value=obs.criado_em.strftime('%d/%m/%Y %H:%M'))
        if i % 2 == 0:
            for c in range(1, 7):
                ws5.cell(row=i, column=c).fill = fill_alt
    ajustar(ws5)

    # ── Gera resposta ────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arq = f'desempenho_{date.today().isoformat()}.xlsx'
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{nome_arq}"'
    return resp


# ─── APIs JSON ────────────────────────────────────────────────────────────────

@login_required
def api_desempenho_turma(request, turma):
    """Retorna JSON com métricas dos alunos de uma turma (para Chart.js)."""
    turma_dec = unquote(turma)
    qs        = _filtrar_resultados(request).filter(turma=turma_dec)

    alunos_dict = defaultdict(list)
    for r in qs.order_by('aluno_nome'):
        alunos_dict[r.aluno_nome].append(r)

    dados = []
    for nome, res in sorted(alunos_dict.items()):
        st = _calcular_stats_aluno(res)
        if st:
            dados.append({
                'nome':             nome,
                'media':            st['media'],
                'classificacao':    st['classificacao'],
                'total_avaliacoes': st['total_avaliacoes'],
            })

    return JsonResponse({'ok': True, 'turma': turma_dec, 'alunos': dados})


@login_required
@require_POST
def api_recalcular(request):
    """
    Recalcula e persiste o cache DesempenhoAluno para todos os alunos visíveis
    pelo perfil do usuário. Retorna JSON com o total processado.
    """
    qs = (
        _filtrar_resultados(request)
        .select_related('avaliacao', 'avaliacao__professor', 'avaliacao__professor__tenant')
        .order_by('turma', 'aluno_nome', 'data_correcao')
    )

    grupos = defaultdict(list)
    for r in qs:
        escola_id = getattr(r.avaliacao.professor, 'tenant_id', 0) or 0
        grupos[(r.aluno_nome, r.turma, escola_id)].append(r)

    total = 0
    for (aluno_nome, turma, escola_id), res in grupos.items():
        st = _calcular_stats_aluno(res)
        if not st:
            continue
        DesempenhoAluno.objects.update_or_create(
            aluno_nome=aluno_nome,
            turma=turma,
            escola_id=escola_id,
            defaults={
                'total_avaliacoes':    st['total_avaliacoes'],
                'media_geral':         st['media'],
                'maior_nota':          st['maior_nota'],
                'menor_nota':          st['menor_nota'],
                'taxa_aprovacao':      st['taxa_aprovacao'],
                'evolucao':            st['evolucao'],
                'classificacao':       st['classificacao'],
                'notas_por_disciplina': st['notas_por_disciplina'],
                'historico_notas':     st['historico'],
            },
        )
        total += 1

    return JsonResponse({'ok': True, 'total_recalculados': total})
